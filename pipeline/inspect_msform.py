"""Print the column headers from the MS Form Excel so we can map them."""
import os
from dotenv import load_dotenv
load_dotenv(".env")
from openpyxl import load_workbook

path = os.getenv("MSFORM_EXCEL_PATH", "").strip()
print(f"File: {path}")
wb = load_workbook(path, read_only=True, data_only=True)
ws = wb.active

# First row = headers
header_row = next(ws.iter_rows(min_row=1, max_row=1, values_only=True))
print(f"Sheet: {ws.title}")
print(f"Columns ({len(header_row)}):")
for i, h in enumerate(header_row):
    print(f"  [{i}] {h!r}")

# Count rows
total = ws.max_row
print()
print(f"Total rows (incl. header): {total}")
