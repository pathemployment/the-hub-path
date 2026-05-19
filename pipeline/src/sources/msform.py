"""Counsellor-submitted jobs from the MS Form Excel response file.

Form columns (Kevin's setup):
  - Id, Start time, Completion time, Email, Name (form metadata)
  - 'Web site url:'           -> the job posting URL
  - 'Label / Note'            -> free-text note from counsellor
  - 'Name of your organization' -> who submitted (ignored)

For each row, we scrape the URL via Firecrawl to pull title/employer/location/
salary from the actual posting page. Results are cached by URL in
data/msform_cache.json so we don't re-scrape the same submission every run.
"""
from __future__ import annotations
import json
import logging
import os
import re
from datetime import datetime
from pathlib import Path

from src.normalize import make_job, infer_region

log = logging.getLogger(__name__)

# Column-header hints (case-insensitive substrings)
COLUMN_HINTS = {
    "url": ["web site url", "website url", "url", "link", "posting"],
    "label": ["label", "note", "description", "details"],
    "submitted_at": ["completion time", "start time", "submitted at"],
}


def fetch(firecrawl_client=None, excel_path: str | None = None,
          project_root: Path | str | None = None) -> list[dict]:
    excel_path = excel_path or os.getenv("MSFORM_EXCEL_PATH", "").strip()
    if not excel_path:
        log.warning("MSFORM_EXCEL_PATH not set — skipping MS Form source")
        return []
    p = Path(excel_path)
    if not p.exists():
        log.warning("MS Form Excel not found at %s — skipping", excel_path)
        return []

    try:
        from openpyxl import load_workbook
    except ImportError:
        log.error("openpyxl not installed")
        return []

    try:
        wb = load_workbook(p, read_only=True, data_only=True)
    except Exception as e:
        log.warning("MS Form Excel could not be opened (%s). Skipping.", type(e).__name__)
        log.warning("  Suggestion: open the file in Excel once to force OneDrive to download it locally.")
        log.warning("  Path: %s", excel_path)
        return []

    ws = wb.active
    rows = list(ws.iter_rows(values_only=True))
    if len(rows) < 2:
        log.info("MS Form: no submissions yet")
        return []

    header = [str(c).strip() if c else "" for c in rows[0]]
    col_index = _detect_columns(header)
    log.info("MS Form columns detected: %s", col_index)

    if "url" not in col_index:
        log.warning("MS Form: couldn't find a URL column — submissions can't be processed")
        return []

    # Cache so we don't re-scrape the same URLs every run.
    if project_root is None:
        project_root = Path(__file__).resolve().parent.parent.parent
    project_root = Path(project_root)
    cache_path = project_root / "data" / "msform_cache.json"
    cache = _load_cache(cache_path)

    jobs: list[dict] = []
    cache_updated = False
    for row in rows[1:]:
        if not row or all(c is None or str(c).strip() == "" for c in row):
            continue

        url = _get(row, col_index, "url")
        if not url or not url.startswith(("http://", "https://")):
            continue
        label = _get(row, col_index, "label")
        submitted_at = _get(row, col_index, "submitted_at")

        if url in cache:
            job = dict(cache[url])  # copy
            # Re-apply region inference each run so cached entries pick up
            # improved logic without needing a re-scrape.
            job["region"] = _infer_region_with_fallbacks(
                location=job.get("location", ""),
                url=url,
                label=label,
                title=job.get("title", ""),
            )
            # Refresh submission date from the form (counsellor may have updated)
            if submitted_at:
                job["date"] = _normalize_date(submitted_at) or job.get("date")
            jobs.append(job)
            continue

        # Not cached — scrape if we can
        if not firecrawl_client:
            log.warning("MS Form: %s — no Firecrawl client to scrape new URL", url)
            continue

        log.info("MS Form: scraping new URL %s", url)
        try:
            scraped = firecrawl_client.scrape(
                url, formats=["markdown", "html"], only_main_content=True
            )
            job = _build_job_from_scrape(
                scraped, url=url, label=label, submitted_at=submitted_at
            )
            jobs.append(job)
            cache[url] = job
            cache_updated = True
        except Exception as e:
            log.warning("  scrape failed for %s: %s", url, e)
            # Fall back to a sparse entry so the submission isn't lost
            jobs.append(_sparse_fallback(url, label, submitted_at))

    if cache_updated:
        _save_cache(cache_path, cache)
    log.info("MS Form: %d submitted jobs (cached + fresh)", len(jobs))
    return jobs


# ---------- helpers ----------

def _detect_columns(header: list[str]) -> dict[str, int]:
    out: dict[str, int] = {}
    lowered = [h.lower() for h in header]
    for field, hints in COLUMN_HINTS.items():
        for i, h in enumerate(lowered):
            if any(hint in h for hint in hints):
                out[field] = i
                break
    return out


def _get(row: tuple, col_index: dict[str, int], field: str) -> str:
    if field not in col_index:
        return ""
    val = row[col_index[field]]
    if val is None:
        return ""
    return str(val).strip()


def _load_cache(path: Path) -> dict:
    if not path.exists():
        return {}
    try:
        return json.loads(path.read_text(encoding="utf-8"))
    except Exception as e:
        log.warning("Failed to read MS Form cache: %s", e)
        return {}


def _save_cache(path: Path, cache: dict) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    path.write_text(json.dumps(cache, ensure_ascii=False, indent=2), encoding="utf-8")


def _build_job_from_scrape(scraped, *, url: str, label: str, submitted_at: str) -> dict:
    md = getattr(scraped, "markdown", "") or ""
    html = getattr(scraped, "html", "") or ""

    # Try Job Bank-specific parser first since most submissions will be JB URLs
    if "jobbank.gc.ca/jobsearch/jobposting/" in url:
        parsed = _parse_jobbank_single(md, html)
    else:
        parsed = _parse_generic(md, html)

    title = parsed.get("title") or label or "(see posting)"
    employer = parsed.get("employer", "")
    location = parsed.get("location", "")
    salary = parsed.get("salary", "Not listed")

    region = _infer_region_with_fallbacks(location=location, url=url, label=label, title=title)

    return make_job(
        title=title,
        employer=employer,
        location=location,
        url=url,
        source="Submitted",
        salary=salary,
        description=label,
        submitted=True,
        region=region,
        date=_normalize_date(submitted_at) or None,
    )


def _infer_region_with_fallbacks(*, location: str, url: str, label: str, title: str) -> str:
    """Try multiple text sources to find a region. Submitted jobs default to
    Hamilton if all signals fail (PATH is headquartered there)."""
    for candidate in (location, url, label, title):
        if not candidate:
            continue
        r = infer_region(candidate)
        if r != "Other":
            return r
    # Counsellor-submitted jobs without any region signal default to Hamilton
    return "Hamilton"


def _sparse_fallback(url: str, label: str, submitted_at: str) -> dict:
    """When we can't scrape, at least record what we know from the form."""
    return make_job(
        title=label or "Submitted job",
        employer="",
        location="",
        url=url,
        source="Submitted",
        salary="Not listed",
        description=label,
        submitted=True,
        region="Other",
        date=_normalize_date(submitted_at) or None,
    )


def _parse_jobbank_single(md: str, html: str) -> dict:
    """Pull title/employer/location/salary from a Job Bank single-posting page."""
    out: dict = {}
    # The page title looks like 'job title - Job Bank' in <title> and in <h1>
    title_m = re.search(r"<title[^>]*>([^<]+?)\s*[-|]\s*Job Bank", html or "", re.I)
    if title_m:
        out["title"] = title_m.group(1).strip()
    elif md:
        # First H1 in markdown
        h1 = re.search(r"^#\s+(.+)$", md, re.M)
        if h1:
            out["title"] = h1.group(1).strip()

    # Employer line "**Employer:** Name" or "Employer\n\nName"
    emp = re.search(r"Employer[:\s]*\n+\s*([^\n]+)", md or "")
    if emp:
        out["employer"] = emp.group(1).strip().lstrip("*").strip()

    # Location: "Location\n\n123 Main St, City (ON)" or "Location: City (ON)"
    loc = re.search(r"Location[:\s]*\n+\s*([^\n]+)", md or "")
    if loc:
        out["location"] = loc.group(1).strip().lstrip("*").strip()

    # Salary: "Salary: $XX.XX hourly" or "Salary\n\n$XX.XX..."
    sal = re.search(r"Salary[:\s]*\n+\s*([^\n]+)", md or "")
    if sal:
        out["salary"] = sal.group(1).strip().lstrip("*").strip()

    return out


def _parse_generic(md: str, html: str) -> dict:
    """Best-effort extraction from any job posting page."""
    out: dict = {}
    # Title from <title> tag (most reliable)
    title_m = re.search(r"<title[^>]*>([^<]+)</title>", html or "", re.I)
    if title_m:
        out["title"] = re.sub(r"\s*[-|]\s*.+$", "", title_m.group(1).strip())
    elif md:
        h1 = re.search(r"^#\s+(.+)$", md, re.M)
        if h1:
            out["title"] = h1.group(1).strip()

    # Labelled fields
    out["employer"] = _find_labelled(md, ["Company", "Employer", "Posted by", "Hiring"]) or ""
    out["location"] = _find_labelled(md, ["Location", "Place of work", "Where", "City"]) or ""
    salary = _find_labelled(md, ["Salary", "Wage", "Pay rate", "Compensation"])
    out["salary"] = salary or "Not listed"
    return out


def _find_labelled(text: str, labels: list[str]) -> str:
    if not text:
        return ""
    for lbl in labels:
        m = re.search(rf"{re.escape(lbl)}\s*[:|]?\s*\n*\s*([^\n]+)", text)
        if m:
            v = m.group(1).strip().lstrip("*").strip()
            if v and len(v) < 200:
                return v
    return ""


def _normalize_date(text: str) -> str:
    if not text:
        return ""
    t = text.strip()
    m = re.search(r"(\d{4})-(\d{2})-(\d{2})", t)
    if m:
        return f"{m.group(1)}-{m.group(2)}-{m.group(3)}"
    # MS Forms sometimes uses M/D/YYYY or MM/DD/YYYY HH:MM:SS
    for fmt in ("%m/%d/%Y", "%m/%d/%Y %H:%M:%S", "%Y-%m-%d %H:%M:%S"):
        try:
            return datetime.strptime(t, fmt).date().isoformat()
        except ValueError:
            continue
    return ""
